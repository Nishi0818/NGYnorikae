from __future__ import annotations

from io import BytesIO
from pathlib import Path
from zipfile import ZipFile
import json

from openpyxl import load_workbook


SOURCE_DIR = Path('/home/ubuntu/nagoya-timetable-source')
OUTPUT_PATH = SOURCE_DIR / 'workbook_structure.json'


def text(value: object) -> str:
    return str(value).strip() if value is not None else ''


def sheet_summary(sheet) -> dict:
    title_cells = [
        text(cell.value)
        for row in sheet.iter_rows(min_row=1, max_row=min(4, sheet.max_row), values_only=False)
        for cell in row
        if text(cell.value)
    ]
    day_markers = []
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            value = text(cell.value)
            if value in {'平日', '土休日'}:
                day_markers.append({'label': value, 'row': cell.row, 'column': cell.column})
    return {
        'name': sheet.title,
        'title_cells': title_cells[:4],
        'day_markers': day_markers,
        'max_row': sheet.max_row,
        'max_column': sheet.max_column,
    }


def archive_summary(archive_path: Path) -> dict:
    workbooks = []
    with ZipFile(archive_path) as archive:
        for item in archive.infolist():
            if not item.filename.lower().endswith('.xlsx'):
                continue
            workbook = load_workbook(BytesIO(archive.read(item)), data_only=True, read_only=True)
            sheets = [sheet_summary(sheet) for sheet in workbook.worksheets]
            workbooks.append({'entry': item.filename, 'sheets': sheets})
            workbook.close()
    return {'archive': archive_path.name, 'workbooks': workbooks}


def main() -> None:
    result = [archive_summary(path) for path in sorted(SOURCE_DIR.glob('*.zip'))]
    OUTPUT_PATH.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding='utf-8')
    print(OUTPUT_PATH)


if __name__ == '__main__':
    main()

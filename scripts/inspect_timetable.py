from __future__ import annotations

from io import BytesIO
from pathlib import Path
from zipfile import ZipFile
import json

from openpyxl import load_workbook


SOURCE_DIR = Path('/home/ubuntu/nagoya-timetable-source')
OUTPUT_PATH = SOURCE_DIR / 'workbook_preview.json'


def preview_archive(archive_path: Path) -> dict:
    with ZipFile(archive_path) as archive:
        workbook_entries = [
            item for item in archive.infolist()
            if item.filename.lower().endswith('.xlsx')
        ]
        previews = []
        for item in workbook_entries[:2]:
            raw = archive.read(item)
            workbook = load_workbook(BytesIO(raw), data_only=True, read_only=True)
            sheet = workbook.active
            rows = []
            for row in sheet.iter_rows(min_row=1, max_row=min(24, sheet.max_row), max_col=min(18, sheet.max_column), values_only=True):
                values = [str(value).strip() if value is not None else '' for value in row]
                if any(values):
                    rows.append(values)
            previews.append(
                {
                    'entry': item.filename,
                    'sheet': sheet.title,
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                    'preview_rows': rows,
                }
            )
            workbook.close()
    return {'archive': archive_path.name, 'workbook_count': len(workbook_entries), 'previews': previews}


def main() -> None:
    archives = sorted(SOURCE_DIR.glob('*.zip'))
    data = [preview_archive(archive) for archive in archives]
    OUTPUT_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    print(OUTPUT_PATH)


if __name__ == '__main__':
    main()

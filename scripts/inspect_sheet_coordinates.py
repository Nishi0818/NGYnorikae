from __future__ import annotations

from io import BytesIO
from pathlib import Path
from zipfile import ZipFile
import json

from openpyxl import load_workbook


ARCHIVE_PATH = Path('/home/ubuntu/nagoya-timetable-source/higashiyama.zip')
OUTPUT_PATH = Path('/home/ubuntu/nagoya-timetable-source/sheet_coordinates.json')


def normalize(value: object) -> str:
    return str(value).strip() if value is not None else ''


def main() -> None:
    with ZipFile(ARCHIVE_PATH) as archive:
        first_entry = next(item for item in archive.infolist() if item.filename.endswith('.xlsx'))
        workbook = load_workbook(BytesIO(archive.read(first_entry)), data_only=True, read_only=True)
        sheet = workbook.active
        rows = []
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=12, max_col=sheet.max_column, values_only=False),
            start=1,
        ):
            cells = [
                {'column': cell.column, 'value': normalize(cell.value)}
                for cell in row
                if normalize(cell.value)
            ]
            rows.append({'row': row_index, 'cells': cells})
        result = {'entry': first_entry.filename, 'sheet': sheet.title, 'rows': rows}
        workbook.close()
    OUTPUT_PATH.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding='utf-8')
    print(OUTPUT_PATH)


if __name__ == '__main__':
    main()

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from zipfile import ZipFile
import json
import re

from openpyxl import load_workbook


SOURCE_DIR = Path('/home/ubuntu/nagoya-timetable-source')
OUTPUT_PATH = Path('/home/ubuntu/nagoya-subway-offline/lib/subway/timetable.generated.ts')
REPORT_PATH = SOURCE_DIR / 'timetable_build_report.json'

ARCHIVES = {
    'higashiyama.zip': {'id': 'higashiyama', 'revision': '2025-03-29'},
    'meijo.zip': {'id': 'meijo', 'revision': '2025-09-29'},
    'meiko.zip': {'id': 'meiko', 'revision': '2025-09-29'},
    'tsurumai.zip': {'id': 'tsurumai', 'revision': '2024-03-16'},
    'sakuradori.zip': {'id': 'sakuradori', 'revision': '2023-09-16'},
    'kamiida.zip': {'id': 'kamiida', 'revision': '2024-03-16'},
}

STATION_RENAMES = {
    '市役所': '名古屋城',
    '伝馬町': '熱田神宮伝馬町',
    '神宮西': '熱田神宮西',
    '中村区役所': '太閤通',
}


def text(value: object) -> str:
    return str(value).strip() if value is not None else ''


def parse_station(header: str) -> str:
    match = re.match(r'^(.*?)駅\s*[（(]', header)
    if not match:
        raise ValueError(f'駅名を表題から読み取れません: {header}')
    return STATION_RENAMES.get(match.group(1).strip(), match.group(1).strip())


def infer_direction(line_id: str, header: str) -> int:
    # 見出しの先頭は「(現在駅)駅」なので、行き先の判定はそれより後ろの文字列だけを見る。
    # 終着駅(例: 名港線の名古屋港)では現在駅名と行き先方面のキーワードが一致してしまい、
    # 判定を見出し全体に対して行うと常に positive 方向と誤判定してしまうため。
    destination_text = header.split('駅', 1)[1] if '駅' in header else header
    if line_id == 'meijo':
        return 1 if '右回り' in destination_text else -1
    positive_destinations = {
        'higashiyama': ('藤が丘',),
        'meiko': ('名古屋港',),
        'tsurumai': ('赤池',),
        'sakuradori': ('徳重',),
        'kamiida': ('上飯田', '小牧', '犬山'),
    }
    return 1 if any(destination in destination_text for destination in positive_destinations[line_id]) else -1


def append_block_times(row: tuple[object, ...], hour_index: int, minute_start_index: int, current_hour: int | None, values: list[int]) -> int | None:
    hour_text = text(row[hour_index]) if len(row) > hour_index else ''
    if re.fullmatch(r'\d{1,2}', hour_text):
        hour = int(hour_text)
        if hour == 0 and current_hour is not None and current_hour >= 23:
            current_hour = 24
        elif current_hour is None or hour != 0:
            current_hour = hour

    if current_hour is None:
        return current_hour

    for index in range(minute_start_index, len(row), 3):
        minute_text = text(row[index])
        if not re.fullmatch(r'\d{1,2}', minute_text):
            continue
        minute = int(minute_text)
        if minute < 60:
            values.append(current_hour * 60 + minute)
    return current_hour


def extract_schedule(sheet) -> tuple[str, list[int], list[int]]:
    header = ''
    for row in sheet.iter_rows(min_row=1, max_row=2, values_only=True):
        for value in row:
            candidate = text(value)
            if '駅' in candidate and ('（' in candidate or '(' in candidate):
                header = candidate
                break
        if header:
            break
    if not header:
        raise ValueError(f'時刻表の表題が見つかりません: {sheet.title}')

    weekday: list[int] = []
    holiday: list[int] = []
    weekday_hour: int | None = None
    holiday_hour: int | None = None
    for row in sheet.iter_rows(min_row=4, values_only=True):
        weekday_hour = append_block_times(row, 1, 4, weekday_hour, weekday)
        holiday_hour = append_block_times(row, 49, 52, holiday_hour, holiday)
    return header, sorted(set(weekday)), sorted(set(holiday))


def main() -> None:
    schedules: dict[str, dict[str, list[int]]] = {}
    report: dict[str, object] = {'lines': {}, 'errors': []}

    for archive_name, metadata in ARCHIVES.items():
        archive_path = SOURCE_DIR / archive_name
        line_id = metadata['id']
        line_count = 0
        with ZipFile(archive_path) as archive:
            entries = [entry for entry in archive.infolist() if entry.filename.lower().endswith('.xlsx')]
            for entry in entries:
                workbook = load_workbook(BytesIO(archive.read(entry)), data_only=True, read_only=True)
                for sheet in workbook.worksheets:
                    try:
                        header, weekday, holiday = extract_schedule(sheet)
                        station = parse_station(header)
                        direction = infer_direction(line_id, header)
                        key = f'{line_id}|{station}|{direction}'
                        schedules[key] = {'weekday': weekday, 'holiday': holiday}
                        line_count += 1
                    except Exception as error:
                        report['errors'].append({'archive': archive_name, 'sheet': sheet.title, 'error': str(error)})
                workbook.close()
        report['lines'][line_id] = {'schedules': line_count, 'revision': metadata['revision']}

    source = (
        '/* 自動生成: scripts/build_offline_timetable.py。直接編集しないでください。 */\n\n'
        f'export const OFFLINE_TIMETABLES = {json.dumps(schedules, ensure_ascii=False, separators=(",", ":"))} as const;\n\n'
        f'export const TIMETABLE_REVISIONS = {json.dumps({value["id"]: value["revision"] for value in ARCHIVES.values()}, ensure_ascii=False)} as const;\n'
    )
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(source, encoding='utf-8')
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'{OUTPUT_PATH} ({len(schedules)} schedules)')
    print(REPORT_PATH)


if __name__ == '__main__':
    main()
